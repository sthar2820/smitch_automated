
import re
from datetime import datetime

# Optional Streamlit import for UI functionality
try:
    import streamlit as st
except ImportError:
    # Create a mock st object for command-line usage
    class MockStreamlit:
        def error(self, msg): print(f"ERROR: {msg}")
        def warning(self, msg): print(f"WARNING: {msg}")
    st = MockStreamlit()

KNOWN_PLANTS = {
    "Bielsko Biala", "Birmingham", "Blatna", "Einbeck", "Forsheda",
    "Olofstrom", "Rotenburg", "Celaya", "Dickson", "Goshen",
    "Kalamazoo", "Saltillo", "Valley City", "Wellington"
}

def extract_weekly_apw(sheet, plant_name=None, part_name=None):
    for row in range(1, 21):  
        for col in range(1, 31): 
            cell = sheet.cell(row=row, column=col).value
            if isinstance(cell, str) and "WEEKLY APW" in cell.upper():
                # Look for numeric value to the right
                for value_col in range(col + 1, min(col + 10, sheet.max_column + 1)):
                    val = sheet.cell(row=row, column=value_col).value
                    numeric_value = extract_numeric_value(val)
                    if numeric_value is not None:
                        metric_name = find_apw_metric_context(sheet, row, value_col)
                        return [{
                            "Category": "EBIT LOSS",
                            "Subcategory": "Weekly APW",
                            "Metric": metric_name,
                            "Value": numeric_value,
                            "Plant": plant_name,
                            "Part Name": part_name
                        }]
    return []

def extract_numeric_value(val):
    if isinstance(val, (int, float)):
        return float(val)
    elif isinstance(val, str):
        clean = re.sub(r'[^\d.\-]', '', val)
        try:
            return float(clean)
        except:
            return None
    return None

def find_apw_metric_context(sheet, row, col):
    offsets = [(-1, 0), (-2, 0), (0, -1), (0, 1)]
    for dr, dc in offsets:
        r, c = row + dr, col + dc
        if r >= 1 and c >= 1:
            val = sheet.cell(row=r, column=c).value
            if isinstance(val, str) and len(val.strip()) > 2:
                text = val.strip().replace('\n', ' ').title()
                if any(word in text.lower() for word in ["quoted", "plex", "actual"]):
                    return text
    return "Weekly APW Value"

def detect_metric_columns(sheet, stop_at_keywords=None):
    if stop_at_keywords is None:
        stop_at_keywords = [
            "demon-strated rate at 100%", "demonstrated rate at 100%",
            "demon-strated rate", "demonstrated rate",
        ]

    metric_cols = []
    headers = {}
    stop_column_found = None

    try:
        for search_row in range(1, min(6, sheet.max_row + 1)):
            temp_cols = []
            temp_headers = {}
            temp_stop_col = None

            for col in range(3, min(sheet.max_column + 1, 20)):
                try:
                    cell = sheet.cell(row=search_row, column=col).value
                    if cell and isinstance(cell, str) and len(cell.strip()) > 1:
                        header_clean = ' '.join(str(cell).strip().split()).lower()
                        temp_headers[col] = header_clean
                        temp_cols.append(col)

                        for stop_keyword in stop_at_keywords:
                            if stop_keyword in header_clean:
                                temp_stop_col = stop_keyword
                                break
                        if temp_stop_col:
                            break
                except:
                    continue

            if temp_stop_col or len(temp_headers) > len(headers):
                headers = temp_headers
                metric_cols = temp_cols
                if temp_stop_col:
                    stop_column_found = temp_stop_col
                    break

        if not metric_cols:
            metric_cols = list(range(3, min(8, sheet.max_column + 1)))
            for col in metric_cols:
                headers[col] = f"column_{chr(64 + col)}".lower()

    except:
        metric_cols = [3, 4, 5, 6]
        headers = {3: "column_c", 4: "column_d", 5: "column_e", 6: "column_f"}

    return metric_cols, headers, stop_column_found


def detect_categories(sheet):
    categories = []
    category_map = {
        'S': 'Sales Price', 'M': 'Material', 'I': 'Investment',
        'T': 'Tooling', 'C': 'Cycle Times', 'H': 'Headcount'
    }

    try:
        for col in range(1, min(4, sheet.max_column + 1)):
            for row in range(1, min(sheet.max_row + 1, 50)):
                try:
                    val = sheet.cell(row=row, column=col).value
                    if not val:
                        continue
                    text = str(val).strip()
                    lines = text.split('\n') if '\n' in text else [text]
                    for line in lines:
                        line_clean = line.strip().upper()
                        if len(line_clean) == 1 and line_clean in category_map:
                            if not any(c['letter'] == line_clean for c in categories):
                                categories.append({
                                    'row': row, 'column': col,
                                    'letter': line_clean,
                                    'name': category_map[line_clean]
                                })
                            break
                except:
                    continue
        categories.sort(key=lambda x: x['row'])
    except Exception as e:
        st.error(f"Error detecting categories: {e}")
        categories = []

    return categories

def detect_plant(sheet):
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            val = sheet.cell(row=row, column=col).value
            if val and isinstance(val, str):
                text = val.strip()
                for plant in KNOWN_PLANTS:
                    if plant.lower() in text.lower():
                        return plant, row
    return None, None

def detect_part_name(sheet, categories):
    try:
        if not categories:
            return None
        first_category_row = categories[0]['row']
        for row in range(first_category_row - 1, 0, -1):  # search upward
            val = sheet.cell(row=row, column=2).value  
            if val and isinstance(val, str):
                val = val.strip()
                # Return first non-empty text that isn't a single letter
                if len(val) > 3 and val.upper() not in {'S', 'M', 'I', 'T', 'C', 'H'}:
                    return val
    except:
        pass
    return None
      

def extract_date(text):
    if not isinstance(text, str):
        return None

    # Enhanced regex to catch both MM/DD/YYYY and YYYY/MM/DD formats
    date_patterns = [
        (r"\b\d{1,2}[/-]\d{1,2}[/-]\d{4}\b", ["%m/%d/%Y", "%m-%d-%Y"]),  # MM/DD/YYYY or MM-DD-YYYY
        (r"\b\d{4}[/-]\d{1,2}[/-]\d{1,2}\b", ["%Y/%m/%d", "%Y-%m-%d"]),  # YYYY/MM/DD or YYYY-MM-DD
    ]
    
    for pattern, formats in date_patterns:
        matches = re.findall(pattern, text)
        for match in matches:
            for date_format in formats:
                try:
                    dt = datetime.strptime(match, date_format)
                    return dt.strftime("%Y-%m-%d")
                except ValueError:
                    continue
    
    # Legacy support for partial dates (keeping existing functionality)
    matches = re.findall(r"\b\d{4}[/-]\d{1,2}(?:[/-]\d{1,2})?\b", text)
    for match in matches:
        try:
            # if re.match(r"\d{1,2}/\d{4}$", match):  # MM/YYYY
            #     dt = datetime.strptime(match, "%m/%Y")
            # elif re.match(r"\d{1,2}/\d{2}$", match):  # MM/YY
            #     dt = datetime.strptime(match, "%m/%y")
            if re.match(r"\d{4}/\d{1,2}/\d{1,2}$", match):  # YYYY/MM/DD
                dt = datetime.strptime(match, "%Y/%m/%d")
            else:
                continue
            return dt.strftime("%Y-%m-%d")
        except:
            continue
    return None



def find_subcategory_column(sheet, categories):
    try:
        if not categories:
            return 3
        category_col = categories[0]['column']
        candidates = [category_col + 1, category_col + 2, 3, 2]
        best_col = category_col + 1
        max_text_cells = 0

        for col in candidates:
            if col < 1 or col > sheet.max_column:
                continue
            text_cells = 0
            for row in range(1, min(30, sheet.max_row + 1)):
                cell = sheet.cell(row=row, column=col).value
                if cell and isinstance(cell, str) and len(cell.strip()) >= 2:
                    text_cells += 1
            if text_cells > max_text_cells:
                max_text_cells = text_cells
                best_col = col
        return best_col
    except:
        return 3

# def extract_ebit_metrics(sheet, plant_name=None, part_name=None, categories=None):
#     extracted = []
#     metric_map = {
#         "quoted cost/pc": "Quoted_Cost",
#         "actual oee cost/pc at plex cost/hr (quote)": "Actual_OEE", 
#         "plex standard cost/pc": "Plex_Cost",
#         "actual oee cost/pc at plex cost/hr (plex)": "Plex_OEE"
#     }
#     allowed_metrics = set(metric_map.values())
    
#     # Step 1: Find the first OH row to identify the EBIT section
#     oh_start_row = None
#     ebit_col = None
    
#     for row in range(1, min(sheet.max_row + 1, 100)):
#         for col in range(1, min(sheet.max_column + 1, 30)):
#             val = sheet.cell(row=row, column=col).value
#             if isinstance(val, str):
#                 val_clean = val.strip().upper()
#                 if val_clean.startswith("OH") and ("$" in val_clean or val_clean == "OH"):
#                     oh_start_row = row
#                     ebit_col = col
#                     break
#         if oh_start_row:
#             break
    
#     if not oh_start_row:
#         return extracted  # No OH section found
    
#     # Step 2: Process rows starting from OH section
#     for row in range(oh_start_row, min(sheet.max_row + 1, oh_start_row + 50)):  # Look within 50 rows
#         val = sheet.cell(row=row, column=ebit_col).value
        
#         if not isinstance(val, str):
#             continue
            
#         val_clean = val.strip().upper()
#         subcategory = None
        
#         # Determine what type of row this is
#         if val_clean.startswith("OH") and ("$" in val_clean or val_clean == "OH"):
#             subcategory = "OH"
#         elif val_clean.startswith("LAB") and ("$" in val_clean or val_clean == "LAB"):
#             subcategory = "LAB"
#         elif "VAR OH TOTAL" in val_clean or "OH TOTAL" in val_clean:
#             subcategory = "OH Total"  # Extract OH summary too
#         elif "LABOR TOTAL" in val_clean or "LAB TOTAL" in val_clean:
#             subcategory = "LAB Total"  # Extract LAB summary too
        
#         # Only process if we have a valid subcategory
#         if not subcategory:
#             continue
            
#         # Get category mapping
#         category = get_category_from_main(categories, row) if categories else "Unknown"
#         seen_metrics = set()
        
#         # Extract values to the right of this row
#         for c in range(ebit_col + 1, min(ebit_col + 15, sheet.max_column + 1)):
#             raw_val = sheet.cell(row=row, column=c).value
#             if raw_val is None:
#                 continue
            
#             try:
#                 if isinstance(raw_val, (int, float)):
#                     value = float(raw_val)
#                 else:
#                     clean_val = str(raw_val).strip().replace("$", "").replace("€", "").replace("£", "").replace(",", "")
#                     if clean_val:
#                         value = float(clean_val)
#                     else:
#                         continue
#             except:
#                 continue

#             # Search upward for metric header
#             metric = None
#             for rh in range(row - 1, max(0, row - 15), -1):  # Look further up for headers
#                 header = sheet.cell(row=rh, column=c).value
#                 if isinstance(header, str) and len(header.strip()) > 3:
#                     header_lower = header.strip().lower()
#                     for k, v in metric_map.items():
#                         if k in header_lower:
#                             metric = v
#                             break
#                     if metric:
#                         break

#             if metric and metric in allowed_metrics and metric not in seen_metrics:
#                 extracted.append({
#                     "Category": category,
#                     "Subcategory": subcategory,
#                     "Metric": metric,
#                     "Value": value,
#                     "Plant": plant_name,
#                     "Part Name": part_name
#                 })
#                 seen_metrics.add(metric)
        
#         # Stop after processing LAB Total (end of section)
#         if subcategory == "LAB Total":
#             break
    
#     return extracted
def extract_ebit_metrics(sheet, plant_name=None, part_name=None, categories=None):
    extracted = []
    metric_map = {
        "quoted cost/pc": "Quoted_Cost",
        "actual oee cost/pc at plex cost/hr (quote)": "Actual_OEE",
        "plex standard cost/pc": "Plex_Cost",
        "actual oee cost/pc at plex cost/hr (plex)": "Plex_OEE"
    }
    allowed_metrics = set(metric_map.values())

    ebit_col = None
    start_row = None

    # Step 1: Locate OH start row and EBIT column
    for row in range(1, sheet.max_row + 1):
        for col in range(1, min(sheet.max_column + 1, 50)):
            val = sheet.cell(row=row, column=col).value
            if isinstance(val, str) and val.strip().upper().startswith("OH"):
                start_row = row
                ebit_col = col
                break
        if ebit_col:
            break

    if not ebit_col:
        return extracted  # OH section not found

    # Step 2: From OH down to LAB Total
    for row in range(start_row, sheet.max_row + 1):
        cell_val = sheet.cell(row=row, column=ebit_col).value
        if not isinstance(cell_val, str):
            continue

        val_clean = cell_val.strip().upper()
        subcategory = None
        if val_clean.startswith("OH") and ("$" in val_clean or val_clean == "OH"):
            subcategory = "OH"
        elif val_clean.startswith("LAB") and ("$" in val_clean or val_clean == "LAB"):
            subcategory = "LAB"
        elif "VAR OH TOTAL" in val_clean or "OH TOTAL" in val_clean:
            subcategory = "OH Total"
        elif "LABOR TOTAL" in val_clean or "LAB TOTAL" in val_clean:
            subcategory = "LAB Total"
        else:
            continue  # Skip non-subcategory rows

        # Stop after LAB Total
        if subcategory == "LAB Total":
            break

        category = get_category_from_main(categories, row) if categories else "Unknown"
        seen_metrics = set()

        for c in range(ebit_col + 1, min(sheet.max_column + 1, ebit_col + 15)):
            raw_val = sheet.cell(row=row, column=c).value
            if raw_val is None:
                continue

            try:
                value = float(str(raw_val).replace("$", "").replace(",", "").strip())
            except:
                continue

            # Search up for metric
            metric = None
            for rh in range(row - 1, max(0, row - 30), -1):
                header = sheet.cell(row=rh, column=c).value
                if isinstance(header, str):
                    lower = header.strip().lower()
                    for k, v in metric_map.items():
                        if k in lower:
                            metric = v
                            break
                    if metric:
                        break

            if metric and metric in allowed_metrics and metric not in seen_metrics:
                extracted.append({
                    "Category": category,
                    "Subcategory": subcategory,
                    "Metric": metric,
                    "Value": value,
                    "Plant": plant_name,
                    "Part Name": part_name
                })
                seen_metrics.add(metric)

    return extracted


def get_category_from_main(categories, target_row):

    if not categories:
        return "Unknown"
    
    # Find the category with the highest row number that's still <= target_row
    best_category = None
    for category in categories:
        if category['row'] <= target_row:
            if best_category is None or category['row'] > best_category['row']:
                best_category = category
    
    return best_category['name'] if best_category else "Unknown"

def extract_smitch_data(sheet, categories, metric_cols, headers, subcategory_col, plant_name=None, part_name=None):
    extracted = []

    if not categories:
        st.warning("No categories found")
        return []

    METRIC_NORMALIZATION = {
        "quoted cost model": "Quoted",
        "quoted": "Quoted",
        "plex standard": "Plex",
        "plex": "Plex",
        "actual performance": "Actual",
        "actual": "Actual",
        "forecasted cost": "Forecasted",
        "forecasted": "Forecasted",
        "demonstrated rate": "Demonstrated",
        "demon-strated": "Demonstrated",
    }

    # Preprocess: Extract date from headers for each metric column
    col_date_map = {}
    for col in metric_cols:
        date_found = None
        for row in range(1, 6):
            cell_val = sheet.cell(row=row, column=col).value
            if isinstance(cell_val, str):
                possible_date = extract_date(cell_val)
                if possible_date:
                    date_found = possible_date
                    break
        col_date_map[col] = date_found

    # Iterate through category rows
    for i in range(len(categories)):
        current = categories[i]
        start_row = current['row']
        end_row = categories[i + 1]['row'] - 1 if i + 1 < len(categories) else min(start_row + 25, sheet.max_row)

        for row in range(start_row, end_row + 1):
            subcat_cell = sheet.cell(row=row, column=subcategory_col).value
            if not subcat_cell:
                continue
            subcat = str(subcat_cell).strip()

            for col in metric_cols:
                val = sheet.cell(row=row, column=col).value
                if not isinstance(val, (int, float)):
                    continue

                # Normalize header
                raw_header = headers.get(col, f"column_{chr(64 + col)}").strip().lower().split('\n')[0]
                if "cm%" in raw_header:
                    continue

                # Clean header for matching
                cleaned_header = re.sub(r'[^a-z\s$\/→-]', '', raw_header.lower()).strip()

                # Match against normalization dict
                matched_key = next((k for k in METRIC_NORMALIZATION if k in cleaned_header), None)

                if matched_key:
                    metric = METRIC_NORMALIZATION[matched_key]
                elif "quoted jph" in cleaned_header:
                    metric = "Quoted_JPH"
                elif "quoted $" in cleaned_header or "quoted $ / piece" in cleaned_header:
                    metric = "Quoted_$"
                elif "actual jph" in cleaned_header:
                    metric = "Actual_JPH"
                elif "actual $" in cleaned_header or "actual $ / piece" in cleaned_header:
                    metric = "Actual_$"
                elif "plex std" in cleaned_header and "jph" in cleaned_header:
                    metric = "Plex_JPH"
                elif "plex std" in cleaned_header and ("$" in cleaned_header or "piece" in cleaned_header):
                    metric = "Plex_$"
                else:
                    metric = raw_header.split()[0].capitalize() if raw_header else f"Col_{col}"

                date_str = col_date_map.get(col)

                entry = {
                    'Category': current['name'],
                    'Subcategory': subcat,
                    'Date': date_str,
                    'Metric': metric,
                    'Value': float(val)
                }
                if plant_name:
                    entry['Plant'] = plant_name
                if part_name:
                    entry['Part Name'] = part_name

                extracted.append(entry)

    return extracted


def extract_smitch_data_from_path(file_path):
    from openpyxl import load_workbook
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    metric_columns, headers, stop_column_found = detect_metric_columns(ws)
    category_rows = detect_categories(ws)
    subcategory_col = find_subcategory_column(ws, category_rows)
    plant_name, plant_row = detect_plant(ws)
    part_name = detect_part_name(ws, category_rows)

    core_data = extract_smitch_data(ws, category_rows, metric_columns, headers, subcategory_col, plant_name, part_name)
    apw_data = extract_weekly_apw(ws, plant_name, part_name)
    ebit_data = extract_ebit_metrics(ws, plant_name, part_name, category_rows)

    return core_data + apw_data + ebit_data
